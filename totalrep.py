#מצג כולל
import custom
import db
import numpy as np
import pandas as pd
import inspect
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, PatternFill,Side
import sqlite3

def totalrep(level="0.2,2000"):

    mydb = db.MYSQLDB()

    levellist = level.split(",")

    cutoffrate = min([float(eachlevel) for eachlevel in levellist])
    cutoffamount = max([float(eachlevel) for eachlevel in levellist])

    conn = sqlite3.connect("dbsave.db")

    query = "SELECT Empid,Empid_mn,Empname,Refdate,Elemtype,Amount as {},Elem,Division,Startdate FROM {} WHERE Elemtype IN ('addition components', 'compulsory deductions', 'voluntary deductions')"

    curdf = pd.read_sql_query(query.format("CurAmount","dfcurr"), conn)
    prevdf = pd.read_sql_query(query.format("PrevAmount","dfprev"), conn)

    REFMONTH = curdf["Refdate"].max()
    PREVMONTH = prevdf["Refdate"].max()

    #curdf = custom.DFCURR.loc[custom.DFCURR["Elemtype"].isin(("addition components","compulsory deductions","voluntary deductions")),["Empid","Empid_mn","Empname","Refdate","Elemtype","Amount","Elem","Division","Startdate"]]   
    #prevdf = custom.DFPREV.loc[custom.DFPREV["Elemtype"].isin(("addition components","compulsory deductions","voluntary deductions")),["Empid","Empid_mn","Empname","Refdate","Elemtype","Amount","Elem","Division","Startdate"]]

    #curdf.rename(columns={"Amount":"CurAmount"},inplace=True)
    #prevdf.rename(columns={"Amount":"PrevAmount"},inplace=True)

    middf = pd.merge(curdf,prevdf,how="outer",on=["Empid","Empname","Refdate","Elemtype","Elem","Division"])

    middf["PrevAmount"] = middf["PrevAmount"].fillna(0) #filling NaN with 0 so that calculations will be correct
    middf["CurAmount"] = middf["CurAmount"].fillna(0)

    curpension = curdf.loc[curdf["Division"] == custom.pensiondepartment,["Empid","Empid_mn"]]
    curpension.set_index("Empid_mn", inplace=True)
    prevnonpension = prevdf.loc[prevdf["Division"] != custom.pensiondepartment,["Empid","Empid_mn"]]
    prevnonpension.set_index("Empid_mn", inplace=True)

    common_index = curpension.index.intersection(prevnonpension.index) #those who were employees in previous and pensioneers in current

    topensionlist = curpension.loc[common_index,"Empid"].unique()

    newemp = curdf.loc[curdf["Startdate"]>=REFMONTH,"Empid"].unique() #new employees

    prevnewemp = prevdf.loc[(prevdf["Startdate"]>=PREVMONTH)&(prevdf["Startdate"]<REFMONTH),"Empid"].unique()

    def calccols(row):
        reslist = []
        reslist.append(row["CurAmount"] if row["Elemtype"] == "addition components" else 0) #middf["GrossCur"]
        reslist.append(row["PrevAmount"] if row["Elemtype"] == "addition components" else 0) #middf["GrossPrev"]
        reslist.append(reslist[0] if row["Refdate"] == REFMONTH else 0) #middf["GrossCurCur"]
        reslist.append(reslist[0] if row["Refdate"] < REFMONTH else 0) #middf["GrossCurRetro"]
        reslist.append(reslist[1] if row["Refdate"] == PREVMONTH else 0) #middf["GrossPrevCur"]
        reslist.append(row["CurAmount"] if row["Elemtype"] == "compulsory deductions" else 0) #middf["TaxesCur"]
        reslist.append(row["CurAmount"] if row["Elemtype"] == "voluntary deductions" else 0) #middf["DeductsCur"]
        reslist.append(np.round(row["CurAmount"] if row["Elemtype"] == "addition components" else -row["CurAmount"],0)) #middf["NetCur"]
        reslist.append(reslist[0]-reslist[1] if row["Elem"] in custom.annualelement else 0) #middf["Annual"]
        reslist.append(reslist[0]-reslist[1] if row["Elem"] in custom.annualvehicle else 0) #middf["Vehicle"]
        reslist.append(row["CurAmount"] if row["Elem"] in custom.gmarheshbon else 0) #middf["Severance"]
        reslist.append(np.round(reslist[0] - reslist[1] - reslist[8] - reslist[9],0)) #middf["Unexplained"]


        return reslist
    #

    middf[["GrossCur","GrossPrev","GrossCurCur","GrossCurRetro","GrossPrevCur","TaxesCur","DeductsCur","NetCur","Annual","Vehicle","Severance","Unexplained"]] = middf.apply(calccols,axis=1,result_type='expand')

    totaldf = middf.groupby(by = ["Empname","Empid"],as_index=False,group_keys=True).sum(["GrossCur","GrossPrev","GrossCurCur","GrossCurRetro","GrossPrevCur","TaxesCur","DeductsCur","NetCur","Annual","Vehicle","Unexplained"])
        
    totaldf.reset_index(inplace=True)

    totaldf.drop(columns=["CurAmount","PrevAmount"],inplace=True)

    totaldf["Pensioneer"] = totaldf.apply(lambda row: 'גמלאי' if middf.loc[middf["Empid"] == row["Empid"],"Division"].max() == 90 else 0, axis=1)

##### Significant Differences Presentation ######

    curlist = totaldf.loc[abs(totaldf["Unexplained"]) >= cutoffamount,"Empid"].unique()

    curstr = str(tuple(curlist.tolist()))

    query1 = "SELECT Empid,Refdate,Elem,Elem_heb,Amount,Empid_mn,Elemtype FROM " 
    query2 = " WHERE Empid IN " + curstr + " AND Refdate >= '{}' AND (Elemtype = 'addition components' OR Elem_heb = '{}' OR Elem = {})".format(PREVMONTH,custom.dayvalue,custom.semelratio)

    curdf = pd.read_sql_query(query1 + "dfcurr" + query2, conn)
    prevdf = pd.read_sql_query(query1 + "dfprev" + query2, conn)

    #curdf =  custom.DFCURR.loc[custom.DFCURR["Empid"].isin(curlist)&(custom.DFCURR["Refdate"] >= custom.PREVMONTH)&((custom.DFCURR["Elemtype"] == "addition components")|(custom.DFCURR["Elem_heb"] == custom.dayvalue)|(custom.DFCURR["Elem"] == custom.semelratio)),["Empid","Refdate","Elem","Elem_heb","Amount","Empid_mn","Elemtype"]]
    #prevdf = custom.DFPREV.loc[custom.DFPREV["Empid"].isin(curlist)&(custom.DFPREV["Refdate"] >= custom.PREVMONTH)&((custom.DFPREV["Elemtype"] == "addition components")|(custom.DFPREV["Elem_heb"] == custom.dayvalue)|(custom.DFPREV["Elem"] == custom.semelratio)),["Empid","Refdate","Elem","Elem_heb","Amount","Empid_mn","Elemtype"]]

    curdf.rename(columns={"Amount":"CurAmount"},inplace=True)
    prevdf.rename(columns={"Amount":"PrevAmount"},inplace=True)

    explaindf = pd.merge(curdf,prevdf,how="outer",on=["Empid","Refdate","Elem","Elem_heb","Empid_mn","Elemtype"])
        
    explaindf["CurAmount"] = explaindf.apply(lambda row: row["CurAmount"] if row["Refdate"] == REFMONTH else 0,axis=1) #nullify current retro payments
        
    groupdf = explaindf.groupby(by=["Empid","Elem","Elem_heb","Empid_mn","Elemtype"],as_index=False,group_keys=True).sum(["PrevAmount","CurAmount"])
        
    #calculate pensionbase for each mn

    for each_mn in groupdf["Empid_mn"].unique():
        
        prevrateS = groupdf.loc[(groupdf['Empid_mn']==each_mn)&(groupdf['Elem'] == custom.semelratio),'PrevAmount']
        prevrate = prevrateS.item() if prevrateS.size > 0 else 0
        prevhourvalS = groupdf.loc[(groupdf['Empid_mn']==each_mn)&(groupdf["Elem_heb"] == custom.dayvalue),'PrevAmount']
        prevhourval = prevhourvalS.item() if prevhourvalS.size > 0 else 0
        currateS = groupdf.loc[(groupdf['Empid_mn']==each_mn)&(groupdf['Elem'] == custom.semelratio),'CurAmount']
        currate = currateS.item() if currateS.size > 0 else 0

        basedict = {"Empid": groupdf.loc[groupdf['Empid_mn']==each_mn,'Empid'].unique(),
                    "Elem": [99999],
                    "Elem_heb": ['בסיס פנסיה מחושב'],
                    "Empid_mn":[each_mn],
                    "Elemtype": ['additional data'], 
                    "PrevAmount": [22*prevhourval*prevrate],
                    "CurAmount": [22*prevhourval*currate]}
                    
        basedf = pd.DataFrame.from_dict(basedict) 

        groupdf = pd.concat([groupdf,basedf], ignore_index=True)
    #

    groupdf["Diff"] = groupdf["CurAmount"] - groupdf["PrevAmount"]


    def significant_diff(row):
            res = 0
            
            if row["Elem"] not in (custom.annualelement + custom.byreport + (custom.semelratio, custom.dayvalue)): #סמלים שלא בבסיס הפנסיה ולא שנתיים
                

                if abs(row['Diff']) >= 100:
                    res = np.round(row['Diff'],0)
    
            #
            
            elif row["Elem"] in custom.byreport: #סמלים בביס הפנסיה שדורשים דווח של חשב שכר

                prevrateS = groupdf.loc[(groupdf['Empid_mn']==row['Empid_mn'])&(groupdf['Elem'] == custom.semelratio),'PrevAmount']
                prevrate = prevrateS.item() if prevrateS.size > 0 else 0
                prevhourvalS = groupdf.loc[(groupdf['Empid_mn']==row['Empid_mn'])&(groupdf["Elem_heb"] == custom.dayvalue),'PrevAmount']
                prevhourval = prevhourvalS.item() if prevhourvalS.size > 0 else 0
                
                if prevrate != 0:
                    currateS = groupdf.loc[(groupdf['Empid_mn']==row['Empid_mn'])&(groupdf['Elem'] == custom.semelratio),'CurAmount']   
                    currate = currateS.item() if currateS.size > 0 else 0
                    updatedamount = 22*prevhourval / prevrate * currate #hypotetical amount based on chages in position rate
                else:
                    updatedamount = 0
                #

                if abs(row['Diff']-updatedamount) >= 100: #if the differences is grater that one that stems from position rate changes, analyze it
                    res = np.round(row['Diff'],0)

                #
            
            return res
    #

    groupdf["Currentdiff"] = groupdf.apply(significant_diff,axis=1)

    curdf = 0

    #if groupdf.loc[groupdf['Currentdiff']!= 0,["Elem_heb","Currentdiff"]].empty:
    #        curdf = pd.DataFrame.from_dict({"Empid":[row["Empid"]],"Elem_heb":["0"],"Currentdiff":[0]})
    #else:
    curdf = groupdf.loc[groupdf['Currentdiff']!= 0,["Empid","Elem_heb","Currentdiff"]]
    #

    curdf.sort_values(by=["Empid"],inplace=True,ignore_index=True)

### Presenting Significant Retros ####    

    prevlist = totaldf.loc[abs(totaldf["GrossCurRetro"])>= cutoffamount,"Empid"].unique()

    query1 = "SELECT Empid,Elem_heb,Elem, SUM(Amount) as Amount FROM dfcurr WHERE Empid IN " + str(tuple(prevlist.tolist())) + f" AND Refdate < '{REFMONTH}' AND Elemtype = 'addition components'" + " GROUP BY Empid,Elem_heb,Elem"

    groupdf = pd.read_sql_query(query1, conn)

    #groupdf.rename(columns={"SUM(Amount)":"Amount"},inplace=True)

    #groupdf = custom.DFCURR.loc[custom.DFCURR["Empid"].isin(prevlist)&(custom.DFCURR["Elemtype"] == "addition components")&(custom.DFCURR["Refdate"] < custom.REFMONTH)].groupby(by = ["Empid","Elem_heb","Elem"],as_index=False,group_keys=True).sum("Amount")
        
    groupdf["Retrodiff"] = groupdf.apply(lambda row: np.round(row["Amount"],0) if abs(row["Amount"]) >= abs(cutoffrate*groupdf['Amount'].sum()) or abs(row["Amount"]) >= cutoffamount/4 else 0 ,axis=1)

    prevdf = groupdf.loc[groupdf['Retrodiff']!= 0,["Empid","Elem_heb","Retrodiff"]]

### All significant differences are combined together ###
    unexpldf = pd.concat([curdf,prevdf],ignore_index=True)


### Explanations Column - Orders, Stops, to Pension and New Employees 

    eomonth = pd.to_datetime(REFMONTH) + pd.DateOffset(months = +1)
    listoflists = mydb.searchorders(REFMONTH,eomonth)
    orderdf = pd.DataFrame(listoflists,columns=['empid','ordercapt','ordertext'])

    orderdf["text"] = orderdf.apply(lambda row: "{} - {}".format(row["ordercapt"] if isinstance(row["ordercapt"],str) else row["ordercapt"].decode('UTF-8'),row["ordertext"] if isinstance(row["ordertext"],str) else row["ordertext"].decode('UTF-8')),axis=1)

    orderdf.drop_duplicates(inplace=True)
        
    totaldf["IniOrder"] = totaldf.apply(lambda row: "; ".join(orderdf.loc[orderdf["empid"] == row["Empid"],"text"].tolist()),axis=1) 

    totaldf["ToPension"] = totaldf.apply(lambda row: "פרש לפנסיה" if row["Empid"] in topensionlist else "", axis=1)

    query1 = "SELECT DISTINCT Empid,Stopname,Stopfrom, Stoptill FROM dfcurr WHERE Stopfrom IS NOT NULL"

    stopdf = pd.read_sql_query(query1, conn)

    #stopdf = custom.DFCURR.loc[(~custom.DFCURR["Stopfrom"].isna()),["Empid","Stopname","Stopfrom","Stoptill"]]
    #stopdf.drop_duplicates(inplace=True)
    #stopdf.reset_index(inplace=True)

    stopdf["text"] = stopdf.apply(lambda row: f"{row['Stopname']} {row['Stopfrom']} - {row['Stoptill']}", axis=1)
    stopdf.drop(columns=["Stopname","Stopfrom","Stoptill"], inplace=True)
    stopdf.rename(columns={"text": "Stop"}, inplace=True)

    totaldf["Stop"] = totaldf.apply(lambda row: stopdf.loc[stopdf["Empid"] == row["Empid"],"Stop"].tolist(),axis=1) 

    def concattext(row):
            res = str(row["IniOrder"])
            
            if len(row["Stop"]) > 0:
                res = res + ' ' + str(row["Stop"][0])
            #
            
            if len(row["ToPension"]) > 0:
                res = res + ' ' + str(row["ToPension"])
            #

            if row["Empid"] in prevnewemp:
                res = res + ' תחילת עבודה בחודש שעבר'
            elif row["Empid"] in newemp:
                res = res + ' תחילת עבודה החודש'
            #
            return res
    #

    totaldf["Order"] = totaldf.apply(concattext,axis=1)


#### Unexplained (Significant differences presentations) merged with totaldf #####

    totaldf = pd.concat([totaldf,unexpldf],ignore_index=True)
    totaldf["NetRank1"] = totaldf.apply(lambda row: totaldf.loc[totaldf["Empid"]==row["Empid"],"NetCur"].max(),axis=1)

    totaldf.sort_values(by=['NetRank1','Empid'], ascending=False, inplace=True, ignore_index=True)
    cols = ["Empid","Empname","Pensioneer","NetCur","GrossCur","GrossPrev","TaxesCur","DeductsCur","Annual","Vehicle","Severance","Unexplained","Elem_heb","Currentdiff","Retrodiff","Order"] 



### Making JSON for Site presentation ####

    jsdf = totaldf.loc[abs(totaldf["NetCur"])>0,["Empid","Empname","Pensioneer","NetCur","GrossCur","GrossPrev","TaxesCur","DeductsCur","Annual","Vehicle","Severance","Unexplained","Order"]]

    jsdf.set_index("Empid",inplace=True)

    jsdict = jsdf.to_dict('index')

    for eachempid in jsdict:
        curjs = curdf.loc[curdf["Empid"] == eachempid,["Elem_heb","Currentdiff"]].to_dict('list')
        prevjs = prevdf.loc[prevdf["Empid"] == eachempid,["Elem_heb","Retrodiff"]].to_dict('list')
            
        jsdict[eachempid]["CurrDiff"]  = curjs if len(curjs["Elem_heb"]) > 0 else ""
        jsdict[eachempid]["RetroDiff"]  = prevjs if len(prevjs["Elem_heb"]) > 0 else ""
    #

#### Final DF ###

    findf = totaldf.loc[:,cols]
    
    findf.set_index("Empid",inplace=True)

    findf.rename(columns={"Empid":"מספר עובד","Empname":"שם","Pensioneer":"גמלאי","NetCur":"נטו","GrossCur":"ברוטו שוטף","GrossPrev":"ברוטו חודש קודם","GrossCurCur":"ברוטו שוטף החודש","GrossCurRetro":"ברוטו רטרו החודש","GrossPrevCur":"ברוטו שוטף חודש שעבר","TaxesCur":"מסים החודש","DeductsCur":"ניכויים החודש","Annual":"תשלומים שנתיים","Vehicle":"תשלומי רכב שנתיים","Severance":"גמר חשבון","Unexplained":"יתרה לא מוסברת","Elem_heb":"סמל שכר","Currentdiff":"הפרשים שוטפים","Retrodiff":"הפרשי רטרו","Order":"הוראה"},inplace=True)


#### Summary Sheet #### 

    wb = load_workbook(filename = custom.xlresfile)
    ws = wb["Sheet"]
    ws.title = "מרכז"

    ws['A1'] = "נטו החודש"
    ws['B1'] = totaldf["NetCur"].sum()

    ws['B3'] = "החודש"
    ws['C3'] = "חודש קודם"
    ws['A4'] = "ברוטו"
    ws['A6'] = "פיצויים"
    ws['A7'] = "תשלומים שנתיים"
    ws['A8'] = "תשלומי רטרו"
    ws['A9'] = "עובדים שלא עבדו בחודש האחר"
    ws['A10'] = "מעבר לפנסיה החודש"
    ws['A11'] = "סכומי הפרשים"

    ws['B4'] = totaldf["GrossCur"].sum()
    ws['C4'] = totaldf["GrossPrev"].sum()
    ws['B4'].border = Border(top=Side(border_style="thin", color="000000"))
    ws['C4'].border = Border(top=Side(border_style="thin", color="000000"))

    piztable = middf.loc[middf["Elem"].isin(custom.pizuim+custom.HodaaMukdemetHayavBL),["PrevAmount","CurAmount"]]
    ws['B6'] = piztable["CurAmount"].sum()
    ws['C6'] = piztable["PrevAmount"].sum()

    anntable = middf.loc[middf["Elem"].isin(custom.annualelement),["PrevAmount","CurAmount"]]
    ws['B7'] = anntable["CurAmount"].sum()
    ws['C7'] = anntable["PrevAmount"].sum()

    ws['B8'] = middf.loc[(middf["Elemtype"]=="addition components")&(middf["Refdate"]<REFMONTH)&(~middf["Elem"].isin(custom.pizuim+custom.HodaaMukdemetHayavBL+custom.annualelement)),"CurAmount"].sum()
    ws['C8'] = middf.loc[(middf["Elemtype"]=="addition components")&(middf["Refdate"]<PREVMONTH)&(~middf["Elem"].isin(custom.pizuim+custom.HodaaMukdemetHayavBL+custom.annualelement)),"PrevAmount"].sum()

    ws['B9'] = totaldf.loc[totaldf["GrossPrev"]==0,"GrossCur"].sum()
    ws['C9'] = totaldf.loc[totaldf["GrossCur"]==0,"GrossPrev"].sum()

    new_pensioneers = middf.loc[middf["Empid"].isin(topensionlist)&(~middf["Elem"].isin(custom.pizuim+custom.HodaaMukdemetHayavBL+custom.annualelement))&(middf["Elemtype"]=="addition components"),["CurAmount","PrevAmount"]]
    ws['B10'] = new_pensioneers["CurAmount"].sum() - new_pensioneers["PrevAmount"].sum()

    ws['B11'] = "=SUM(B6:B10)"
    ws['B11'].border = Border(top=Side(border_style="thin", color="000000"))

    ws['C11'] = "=SUM(C6:C10)"
    ws['C11'].border = Border(top=Side(border_style="thin", color="000000"))
    ws['B12'] = "=B4-B11"
    ws['C12'] = "=C4-C11"
    ws['D13'] = "=B12-C12"

    for cells in ws["B1:D13"]:
        for eachcell in cells:
            eachcell.number_format = "#,##0;-#,##0;0"
        #
    #

    wb.save(custom.xlresfile)

    with pd.ExcelWriter(custom.xlresfile, mode="a",if_sheet_exists='replace') as writer:
        findf.to_excel(writer,sheet_name="נטו",float_format="%.2f",index=True)
    #
     

    conn.close()
    return [inspect.stack()[0][3],jsdict,"טבלת השוואה מוכנה"]
#