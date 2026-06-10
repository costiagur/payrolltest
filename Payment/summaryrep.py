import custom
import pandas as pd
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, PatternFill,Side

def summaryrep():

    conn = sqlite3.connect(custom.dbsave)

    query = """
    select sum(dfcurr.Amount) as Curr, prevdf.Prev as Prev, sum(dfcurr.Amount) - prevdf.Prev as Diff, "נטו" as Title,0 as Indx
    FROM dfcurr
    LEFT JOIN (SELECT SUM(dfprev.Amount) as Prev FROM dfprev WHERE dfprev.Elem = 91096) as prevdf
    WHERE dfcurr.Elem = 91096 
    UNION
    SELECT "","","","",1
    UNION
    select sum(dfcurr.Amount) as Curr, prevdf.Prev, sum(dfcurr.Amount) - prevdf.Prev as Diff, "ברוטו", 2
    FROM dfcurr
    LEFT JOIN (SELECT SUM(dfprev.Amount) as Prev FROM dfprev WHERE dfprev.ElemType = 'addition components') as prevdf
    WHERE dfcurr.ElemType = 'addition components' 
    UNION
    SELECT SUM(dfcurr.Amount) as Curr, prevdf.Prev, SUM(dfcurr.Amount) - prevdf.Prev as Diff, "פיצויים", 3
    FROM dfcurr
    LEFT JOIN (SELECT SUM(dfprev.Amount) as Prev FROM dfprev WHERE dfprev.Elem IN ("1262","258","261","271","1261","1263","257")) as prevdf
    WHERE dfcurr.Elem IN ("1262","258","261","271","1261","1263","257") AND  dfcurr.Refdate = (SELECT MAX(dfcurr.Refdate) FROM dfcurr)
    UNION
    SELECT SUM(dfcurr.Amount) as Curr, prevdf.Prev, SUM(dfcurr.Amount) - prevdf.Prev as Diff, "חד_שנתי", 4
    FROM dfcurr
    LEFT JOIN (SELECT SUM(dfprev.Amount) as Prev FROM dfprev WHERE dfprev.Elem IN ("2276","2278","290","291","292","295","2151","4737","5831","270","585","5875", "501", "502", "278")) as prevdf
    WHERE dfcurr.Elem IN ("2276","2278","290","291","292","295","2151","4737","5831","270","585","5875", "501", "502", "278") AND  dfcurr.Refdate = (SELECT MAX(dfcurr.Refdate) FROM dfcurr)
    UNION
    SELECT SUM(dfcurr.Amount) as Curr, prevdf.Prev, SUM(dfcurr.Amount) - prevdf.Prev as Diff, "מילואים", 5
    FROM dfcurr
    LEFT JOIN (SELECT SUM(dfprev.Amount) as Prev FROM dfprev WHERE dfprev.Elem IN ("296","297","5780")) as prevdf
    WHERE dfcurr.Elem IN ("296","297","5780") AND dfcurr.Refdate = (SELECT MAX(dfcurr.Refdate) FROM dfcurr)
    UNION
    SELECT SUM(dfcurr.Amount) as Curr, prevdf.Prev, SUM(dfcurr.Amount) - prevdf.Prev as Diff, "עבודה_חופשות_חינוך", 6
    FROM dfcurr
    LEFT JOIN (SELECT SUM(dfprev.Amount) as Prev FROM dfprev WHERE dfprev.Elem IN ("450","451")) as prevdf
    WHERE dfcurr.Elem IN ("450","451") AND dfcurr.Refdate = (SELECT MAX(dfcurr.Refdate) FROM dfcurr)
    UNION
    SELECT SUM(dfcurr.Amount) as Curr, prevdf.Prev, SUM(dfcurr.Amount) - prevdf.Prev AS Diff, "רטרו", 7
    FROM dfcurr
    LEFT JOIN (SELECT SUM(dfprev.Amount) as Prev FROM dfprev WHERE dfprev.Refdate < (SELECT MAX(dfprev.Refdate) FROM dfprev) AND dfprev.ElemType = 'addition components') AS prevdf 
    WHERE dfcurr.Refdate < (SELECT MAX(dfcurr.Refdate) FROM dfcurr) AND dfcurr.ElemType = 'addition components'
    UNION
    SELECT SUM(dfcurr.Amount) as Curr, 0 AS Prev,SUM(dfcurr.Amount) AS Diff, "עובדים_חדשים", 8
    FROM dfcurr
    WHERE dfcurr.ElemType = 'addition components' AND dfcurr.Empid_mn NOT IN (SELECT DISTINCT dfprev.Empid_mn FROM dfprev) AND dfcurr.Refdate = (SELECT MAX(dfcurr.Refdate) FROM dfcurr)
    UNION 
    SELECT 0 as Curr, SUM(dfprev.Amount) AS Prev, -SUM(dfprev.Amount) AS Diff, "לא_עבדו_החודש", 9
    FROM dfprev
    WHERE dfprev.ElemType = 'addition components' AND dfprev.Empid_mn NOT IN (SELECT DISTINCT dfcurr.Empid_mn FROM dfcurr) AND dfprev.Elem NOT IN ("1262","258","261","271","1261","1263","257")
    AND dfprev.Refdate = (SELECT MAX(dfprev.Refdate) FROM dfprev)
    UNION
    SELECT SUM(dfcurr.Amount) AS Curr, SUM(dfprev.Amount) AS Prev, SUM(dfcurr.Amount)-SUM(dfprev.Amount) AS Diff,"גמלאים_חדשים", 10
    FROM dfcurr
    JOIN dfprev ON dfprev.Empid = dfcurr.Empid AND dfprev.ElemType = 'addition components' AND dfprev.Division <> 90 AND dfprev.Elem NOT IN ("1262","258","261","271","1261","1263","257") AND dfprev.Refdate = (SELECT MAX(dfprev.Refdate) FROM dfprev)
    WHERE dfcurr.ElemType = 'addition components' AND dfcurr.Elem NOT IN ("1262","258","261","271","1261","1263","257") AND dfcurr.Division = 90 AND dfcurr.Refdate = (SELECT MAX(dfcurr.Refdate) FROM dfcurr) 
    UNION 
    SELECT SUM(dfcurr.Amount) AS Curr, SUM(dfprev.Amount) AS Prev, SUM(dfcurr.Amount) - SUM(dfprev.Amount) AS Diff, "ניכוי_פנסיוני", 11
    FROM dfcurr
    JOIN dfprev ON dfprev.Empid = dfcurr.Empid AND dfprev.Elem IN (9150, 9151) AND dfprev.Refdate = (SELECT MAX(dfprev.Refdate) FROM dfprev)
    WHERE dfcurr.Elem IN (9150, 9151) AND dfcurr.Refdate = (SELECT MAX(dfcurr.Refdate) FROM dfcurr)
    UNION
    SELECT SUM(dfcurr.Amount) AS Curr,prevdf.Prev,  SUM(dfcurr.Amount) - prevdf.Prev AS Diff, "ביטוח_ורישיון", 12
    FROM dfcurr
    LEFT JOIN (SELECT SUM(dfprev.Amount) as Prev FROM dfprev WHERE dfprev.Elem IN ("143","7143","149","7149","150","7150") AND dfprev.Refdate = (SELECT MAX(dfprev.Refdate) FROM dfprev)) as prevdf
    WHERE dfcurr.Elem IN ("143","7143","149","7149","150","7150") AND dfcurr.Refdate = (SELECT MAX(dfcurr.Refdate) FROM dfcurr)
    UNION
    SELECT SUM(dfcurr.Amount) AS Curr, SUM(dfprev.Amount) AS Prev, SUM(dfcurr.Amount) - SUM(dfprev.Amount) AS Diff, "בסיס_פנסיה_מול_בסיס_פנסיה", 13
    FROM dfcurr
    JOIN dfprev ON dfprev.Empid = dfcurr.Empid AND dfprev.Elem = 91025 AND dfprev.Refdate = (SELECT MAX(dfprev.Refdate) FROM dfprev)
    WHERE dfcurr.Elem = 91025 AND dfcurr.Refdate = (SELECT MAX(dfcurr.Refdate) FROM dfcurr)
    """
    totaldf = pd.read_sql_query(query, conn)
    pd.set_option('display.float_format', lambda x: '%.0f' % x)

    totaldf.set_index(keys="Indx",drop=True, inplace=True)
    totaldf.sort_index(inplace=True)
    
    #### Summary Sheet #### 

    with pd.ExcelWriter(custom.xlresfile, mode="a") as writer:
            totaldf.to_excel(writer,sheet_name="מרכז",index=False,header=["החודש","חודש שעבר","הפרש","סעיף"])

    wb = load_workbook(filename = custom.xlresfile)
    
    ws = wb["Sheet"]
    wb.remove(ws)
    
    ws = wb["מרכז"]
    wb.move_sheet("מרכז",-20)  
    ws['D15'] = "סכומי הפרשים"
    ws['C15'] = "=SUM(C4:B14)"
    ws['C15'].border = Border(top=Side(border_style="thin", color="000000"))
    ws['C16'].border = Border(top=Side(border_style="thin", color="000000"))
    ws['C16'] = "=C3-C15"

    for cells in ws["A2:C16"]:
        for eachcell in cells:
            eachcell.number_format = "#,##0;-#,##0;0"
        #
    #
#
