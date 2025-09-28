import pandas as pd
import numpy as np
import custom
import os
import re
from openpyxl import Workbook, load_workbook
import db
import datetime
import mysql.connector
from sklearn.neighbors import KNeighborsRegressor
from io import BytesIO

def analysis13m(filesdict,expectedplus):

    Nneighbors = 3
    KNR = KNeighborsRegressor(n_neighbors=Nneighbors, weights='distance', algorithm='auto', metric='minkowski')

    mydb = db.MYSQLDB()

    pd.set_option('display.float_format', lambda x: '%.2f' % x)

    
    # פריסת דוח חזותי לפי תאריכי ערך ל- 13 חודשים 
   
    if len(filesdict) > 0:

        byte_data = filesdict['hazuti13m'][1]
        datelist = []
        with BytesIO(byte_data) as f:
            dateline = f.readlines()[3].decode('cp1255').split("\t")
            for each in dateline:
                res = re.findall(r'\d{2}/\d{2}/\d{4}',each)
                if res and res[0] not in datelist:
                    datelist.append(res[0])
                #
            #
        #

        print(datelist)

        hdf = pd.read_csv(BytesIO(byte_data),sep='\t',header=0,encoding="cp1255",na_filter=True,skiprows=4,skip_blank_lines=True,parse_dates=['ת.ת. עבודה'],dayfirst=True) 
        hdf = hdf.iloc[1:,2:46]
        hdf.rename(columns={"שם עובד":"Empname","מספר עובד":"Empid","ת.ז.":"id","מ.נ.":"mn","דרוג":"Dirug","דרגה":"darga","ת.ת. עבודה":"Startdate","וותק":"vetek","אגף":"Division","שם אגף":"Divisionname","מחלקה":"mahlaka","שם מחלקה":"mahlakaname","סוג רכיב":"Elemtype_heb","שם רכיב":"Elem_heb","הפסקה מ":"Stopfrom","הפסקה עד":"Stoptill","סמל הפסקה":"Stopcode","שם הפסקה":"Stopname"}, inplace=True)

        hdf["Empid"] = hdf["Empid"].astype(int)
        hdf["mn"] = hdf["mn"].astype(int)
        hdf["Division"] = hdf["Division"].astype(int)

        hdf["Elem"] = hdf["Elem_heb"].str.extract(r'^(\d+|עלות)\s-*') #extract element num or incase of alut which is without number, seti it to alut. Therefore it is str.
        hdf["Rank"] = hdf["Dirug"].str.extract(r'(\d+)') #extract rank number
        hdf["Empid_mn"] = hdf[["Empid","mn"]].apply(lambda a: "{}_{}".format(a["Empid"],a["mn"]), axis=1)

        fromconv = ["מספר ותאור רכיבי תוספות","מספר ותאור רכיבי ניכויי חובה","מספר ותאור רכיבי ניכויי רשות","מספר ותאור רכיבי הפרשות","נתונים נוספים","מספר ותאור  רכיבי זקיפות הטבה"]
        toconv = ["addition components","compulsory deductions","voluntary deductions","provision components","additional data","benefit charge components"]

        hdf["Elemtype"] = hdf["Elemtype_heb"]
        hdf["Elemtype"] = hdf["Elemtype"].replace(to_replace = fromconv, value=toconv)

        hdf.drop(columns=["id","Divisionname","mahlaka","mahlakaname","Elemtype_heb","Elem_heb","Stopfrom","Stoptill","Stopcode","Stopname"],inplace=True)

        droplist = []
        for i in range(1,13,1):
            droplist.append("כמות.{}".format(i))
        #

        hdf.drop(columns=droplist+["כמות"],inplace=True)

        hdf = hdf.fillna(0)

        middf = hdf.loc[:,"סכום":"סכום.12"]
        numdf = pd.DataFrame({"Amount":[0],"Refdate":[0]})
            
        for ind,eachcol in enumerate(middf.columns):
            newdf = pd.DataFrame(middf[eachcol])
            newdf.rename(columns={eachcol:"Amount"},inplace=True)
            newdf["Refdate"] = pd.to_datetime(datelist[ind],dayfirst=True)
            numdf = pd.concat([numdf,newdf])
        #
            
        numdf.drop(index=[0],inplace=True)
        numdf.reset_index(inplace=True)
        numdf.drop(columns=["index"],inplace=True)

        droplist = ["סכום"]
        for i in range(1,13,1):
            droplist.append("סכום.{}".format(i))
        #

        hdf.drop(columns=droplist,inplace=True)

        hdf = pd.concat([hdf] * 13, ignore_index=True)

        histdf = pd.concat([hdf,numdf],axis=1)

        pd.to_pickle(histdf,r'drafts\prevyear.pkl')
    #
    else:
        histdf = pd.read_pickle(r'drafts\prevyear.pkl')
    #

 
    #Group Historical Data for Analysis

    cols = ["Empid","Empname","Elemtype","Elem","Amount","Rank","Refdate","mn","Empid_mn","Division","darga"]

    def applystack(row):
            res = []
            res = res + [row["Amount"] if row["Elem"] == custom.semelratio else 0]
            res = res + [row["Amount"] if row["Elem"] == custom.pensionbasesemel else 0]
            res = res + [row["Amount"] if row["Elem"] in custom.annualelement else 0]
            res = res + [row["Amount"] if row["Elem"] in [item for item in custom.annualvehicle if item not in custom.gilum] else 0]
            res = res + [row["Amount"] if row["Elem"] in (custom.pizuim + custom.HodaaMukdemetHayavBL + ("257",)) else 0]
            res = res + [row["Amount"] if row["Elem"] in custom.miluim else 0]
            res = res + [row["Amount"] if row["Elem"] in custom.expense else 0]
            res = res + [row["Amount"] if row["Elem"] in custom.gilum else 0]
            res = res + [row["Amount"] if row["Elem"] in custom.deduct else 0] #annual elements in basis

            return res
    #

    if "peruserdata.pkl" not in os.listdir('drafts'):

        middf = histdf.loc[((histdf["Elemtype"] == "addition components")|(histdf["Elem"] == custom.pensionbasesemel)|(histdf["Elem"] == custom.semelratio)),cols]

        middf[["90148","basis","annual","annualvehicle","pizuim","miluim","expense","gilum","deduct"]] = middf.apply(applystack,axis=1,result_type='expand')

        middf.loc[(middf["Elem"] == custom.pensionbasesemel)|(middf["Elem"] == custom.semelratio),"Amount"] = 0

        Xdf = middf.groupby(by=["Empid","Empname","mn","Empid_mn","Rank","Refdate","Division","darga"],as_index=False,group_keys=True).sum(["Amount","basis","annual","annualvehicle","pizuim","miluim","expense","90148","gilum","deduct"])

        Xdf["basis"] = Xdf["basis"] - Xdf["deduct"]

        Xdf["extrawork"] = np.round(Xdf["Amount"] - Xdf["basis"] - Xdf["annual"] - Xdf["annualvehicle"] - Xdf["pizuim"] - Xdf["miluim"] - Xdf["expense"] - Xdf["gilum"])

        Xdf['Rank'] = Xdf['Rank'].fillna(0)

        Xdf["Rank"] = Xdf["Rank"].astype(int)

        pd.to_pickle(Xdf,r'drafts\peruserdata.pkl')
    #
    else:
        Xdf = pd.read_pickle(r'drafts\peruserdata.pkl')
    #

    # Prepare current period data for analysis
    DFCURR = pd.read_pickle('drafts/dfcurr.pkl')
    REFMONTH = DFCURR["Refdate"].max()

    expectedplus = float(expectedplus)/100 if expectedplus else 0.0

    middf = DFCURR.loc[(DFCURR["Refdate"]==REFMONTH)&((DFCURR["Elemtype"] == "addition components")|(DFCURR["Elem"] == custom.pensionbasesemel)|(DFCURR["Elem"] == custom.semelratio)),cols]

    middf[["90148","basis","annual","annualvehicle","pizuim","miluim","expense","gilum","deduct"]] = middf.apply(applystack,axis=1,result_type='expand')

    middf.loc[(middf["Elem"] == custom.pensionbasesemel)|(middf["Elem"] == custom.semelratio),"Amount"] = 0

    Rdf = middf.groupby(by=["Empid","Empname","mn","Empid_mn","Rank","Refdate","Division","darga"],as_index=False,group_keys=True).sum({"Amount","basis","expense","annual","annualvehicle","pizuim","miluim","90148","gilum","deduct"})

    Rdf["basis"] = Rdf["basis"] - Rdf["deduct"]

    Rdf["extrawork"] = np.round(Rdf["Amount"] - Rdf["basis"] - Rdf["annual"] - Rdf["annualvehicle"] - Rdf["pizuim"] - Rdf["miluim"] - Rdf["expense"] - Rdf["gilum"],2)

    Rdf["Rank"] = Rdf["Rank"].astype(int)

    def applytest(row):
        res = []
        res.append(0) # 0 signal of basis deviation
        res.append(0) # 1 signal of extrawork deviation
        res.append(0) # 2 signal of expense deviation
        res.append(0) # 3 basismean
        res.append(0) # 4 basisstd
        res.append(0) # 5 extraworkmean
        res.append(0) # 6 extraworkstd
        res.append(0) # 7 expensemean
        res.append(0) # 8 expensestd
        res.append(0) # 9 wagerate

        
        wagerate = Xdf.loc[(Xdf["Empid_mn"] == row["Empid_mn"])&(Xdf["Division"] == 90 if row['Division'] == 90 else Xdf["Division"] != 90)&(Xdf["90148"]>0),"90148"].values
    
        if len(wagerate) > 0:
            res[9] = np.mean(wagerate)
        else:
            res[9] = 'N/A'
        #

        basis = Xdf.loc[(Xdf["Empid_mn"] == row["Empid_mn"])&(Xdf["Division"] == 90 if row['Division'] == 90 else Xdf["Division"] != 90)&(Xdf["basis"]>0),["basis","Refdate","90148"]]
        extrawork = Xdf.loc[(Xdf["Empid_mn"] == row["Empid_mn"])&(Xdf["Division"] == 90 if row['Division'] == 90 else Xdf["Division"] != 90)&(Xdf["extrawork"]>0),["Refdate","extrawork","90148"]]
        expense = Xdf.loc[(Xdf["Empid_mn"] == row["Empid_mn"])&(Xdf["Division"] == 90 if row['Division'] == 90 else Xdf["Division"] != 90)&(Xdf["expense"]>0),["Refdate","expense","90148"]]
        
        if np.count_nonzero(basis["basis"] > 200) > 3: #if there are more than 3 basis values greater than 200, then we can calculate the std and mean

            currdf = pd.DataFrame({"Days":[0],"90148":[row["90148"]]})

            if  len(basis["basis"]) > Nneighbors:
                basisstd = np.std(basis["basis"])
                basis["Days"] = (REFMONTH - basis["Refdate"]).dt.days
                KNR.fit(basis[["Days","90148"]],basis["basis"].values)            
                basismean = KNR.predict(currdf)

                if (row["basis"]-basismean[0]*(1+expectedplus)) > basisstd*2 and abs(row["basis"]-basismean[0]*(1+expectedplus)) > 200:
                    res[0] = 1
                    res[3] = basismean[0]
                    res[4] = basisstd
                #
            #

            if  len(extrawork["extrawork"]) > Nneighbors:
                extraworkstd = np.std(extrawork["extrawork"])
                extrawork["Days"] = (REFMONTH - extrawork["Refdate"]).dt.days
                KNR.fit(extrawork[["Days","90148"]],extrawork["extrawork"].values) 
                extraworkmean = KNR.predict(currdf)

                if (row["extrawork"]-extraworkmean[0]*(1+expectedplus)) > extraworkstd*2 and abs(row["extrawork"]-extraworkmean*(1+expectedplus)) > 300:
                    res[1] = 1       
                    res[5] = extraworkmean[0]
                    res[6] = extraworkstd
                #
            #

            if  len(expense["expense"]) > Nneighbors:
                expensestd = np.std(expense["expense"])
                expense["Days"] = (REFMONTH - expense["Refdate"]).dt.days
                KNR.fit(expense[["Days","90148"]],expense["expense"].values)           
                expensemean = KNR.predict(currdf)

                if (row["expense"]-expensemean[0]) > expensestd*2 and abs(row["expense"]-expensemean[0]) > 200:
                    res[2] = 1
                    res[7] = expensemean[0]
                    res[8] = expensestd
                #
            #
    
        return res
    #

    Rdf[["sig_basis","sig_extrawork","sig_expense","basismean","basisstd","extraworkmean","extraworkstd","expensemean","expensestd","wageratemean"]] = Rdf.apply(applytest,axis=1,result_type='expand')

    resdf = Rdf.loc[(((Rdf["sig_basis"] == 1)|(Rdf["sig_extrawork"] == 1)|(Rdf["sig_expense"] == 1)))]
    resdf = resdf.drop(columns=["Empid_mn","Refdate"])


    eomonth = pd.to_datetime(REFMONTH) + pd.DateOffset(months = +1)
    preveomonth = pd.to_datetime(REFMONTH) + pd.DateOffset(months = -1)
    listoflists = mydb.searchorders(preveomonth,eomonth)
    orderdf = pd.DataFrame(listoflists,columns=['empid','ordercapt','ordertext'])

    orderdf["text"] = orderdf.apply(lambda row: "{} - {}".format(row["ordercapt"] if isinstance(row["ordercapt"],str) else row["ordercapt"].decode('UTF-8'),row["ordertext"] if isinstance(row["ordertext"],str) else row["ordertext"].decode('UTF-8')),axis=1)

    orderdf.drop_duplicates(inplace=True)
        
    resdf["Order"] = resdf.apply(lambda row: "; ".join(orderdf.loc[orderdf["empid"] == row["Empid"],"text"].tolist()),axis=1) 

    def dargachange(row):
        previous = Xdf.loc[(Xdf["Empid"] == row["Empid"])&(Xdf["mn"] == row["mn"]),"darga"].unique()[0]
        present = row["darga"]

        if present != previous:
            return "חדש {} היה {}".format(present,previous)
        #
        else:
            return ""
        #
    #

    resdf["dargaChange"] = resdf.apply(dargachange,axis=1)

    filename = "drafts/outliers.xlsx"

    resdf[["Empid","Empname","mn","Rank","Division","Amount","sig_basis","sig_extrawork","sig_expense","90148","wageratemean","basis","extrawork","expense","basismean","basisstd","extraworkmean","extraworkstd","expensemean","expensestd","annual","annualvehicle","pizuim","miluim","gilum","Order","dargaChange"]].to_excel(filename,index=False)

    wb = load_workbook(filename)
    ws = wb.active
    ws["A1"] = "מס_עובד"
    ws["B1"] = "שם"
    ws["C1"] = "מנ"
    ws["D1"] = "דרגה"
    ws["E1"] = "מחלקה"
    ws["F1"] = "ברוטו"
    ws["G1"] = "חריגה בסיס הפנסיה"
    ws["H1"] = "חריגה בעבודה נוספת"
    ws["I1"] = "חריגה בהוצאות"
    ws["J1"] = "שיעור משרה"
    ws["K1"] = "ממוצע שיעור משרה"
    ws["L1"] = "בסיס הפנסיה"
    ws["M1"] = "עבודה נוספת"
    ws["N1"] = "הוצאות"
    ws["O1"] = "ממוצע בסיס הפנסיה"
    ws["P1"] = "סטיית תקן בסיס הפנסיה"
    ws["Q1"] = "ממוצע עבודה נוספת"
    ws["R1"] = "סטיית תקן עבודה נוספת"
    ws["S1"] = "ממוצע הוצאות"
    ws["T1"] = "סטיית תקן הוצאות"
    ws["U1"] = "תשלום שנתי"
    ws["V1"] = "תשלום שנתי רכב"
    ws["W1"] = "פיצויים"
    ws["X1"] = "מילואים"
    ws["Y1"] = "גילום"
    ws["Z1"] = "הוראות"
    ws["AA1"] = "שינוי דרגה"
    ws["AB1"] = "הערות"
    
    wb.save(filename)

    return open(filename,'rb').read()



