#מס הכנסה גבוה
import numpy as np
import pandas as pd
import custom
import inspect
import sqlite3
from openpyxl import Workbook, load_workbook

def hightax(level = ""):
    
    pd.set_option('display.float_format', lambda x: '%.2f' % x)

    conn = sqlite3.connect(custom.dbsave)
    cur = conn.cursor()

    query = """SELECT dfcurr.Empid, dfcurr.Empname, dfcurr.Elemtype, dfcurr.Elem_heb, dfcurr.Elem, dfcurr.Rank, dfcurr.Division, SUM(dfcurr.Amount) as CurAmount
      FROM dfcurr 
      WHERE dfcurr.Elemtype IN ('addition components', 'benefit charge components', 'compulsory deductions')
      GROUP BY dfcurr.Empid, dfcurr.Elemtype, dfcurr.Elem_heb, dfcurr.Elem, dfcurr.Rank, dfcurr.Division;"""

    middf = pd.read_sql_query(query, conn)

    query = """SELECT dfprev.Empid, dfprev.Empname, dfprev.Elemtype, dfprev.Elem_heb, dfprev.Elem, dfprev.Rank, dfprev.Division, SUM(dfprev.Amount) as PrevAmount
      FROM dfprev 
      WHERE dfprev.Elemtype IN ('addition components', 'benefit charge components', 'compulsory deductions')
      GROUP BY dfprev.Empid, dfprev.Elemtype, dfprev.Elem_heb, dfprev.Elem, dfprev.Rank, dfprev.Division;"""

    prevdf = pd.read_sql_query(query, conn)

    middf = pd.concat([middf,prevdf],axis=0,ignore_index=True)

    middf.fillna(0,inplace=True)

    def apply1(row):
            reslist = []

            reslist.append(row["PrevAmount"] if row["Elem"] in custom.elembtl else 0) #btlPrev
            reslist.append(row["CurAmount"] if row["Elem"] in custom.elembtl else 0) #btlCur
            reslist.append(row["PrevAmount"] if row["Elem"] == custom.elemtax else 0) #taxPrev
            reslist.append(row["CurAmount"] if row["Elem"] == custom.elemtax else 0) #taxCur
            reslist.append(row["CurAmount"] if row["Elemtype"] in ("addition components","benefit charge components") else 0) #Current Gross Amounts
            reslist.append(row["PrevAmount"] if row["Elemtype"] in ("addition components","benefit charge components") else 0) #Previous Gross Amounts

            return reslist
    #

    middf[["btlPrev","btlCur","taxPrev","taxCur","CurGross","PrevGross"]] =  middf.apply(apply1,axis=1,result_type='expand')

    middf.drop(columns=["CurAmount","PrevAmount"],inplace=True)

    groupdf = middf.groupby(by = ["Empname","Empid","Division"],as_index=False,group_keys=True).agg({"CurGross":'sum',"PrevGross":'sum',"btlCur":'sum',"btlPrev":'sum',"taxCur":'sum',"taxPrev":'sum'})
    
    groupdf.reset_index(drop=True,inplace=True)

    def apply2(row):
            reslist = []

            reslist.append(row["btlCur"] / row["CurGross"] if row["CurGross"] != 0 else np.nan) #btlrateCur
            reslist.append(row["taxCur"] / row["CurGross"] if row["CurGross"] != 0 else np.nan) #taxrateCur

            return reslist
        #

    groupdf[["btlrateCur","taxrateCur"]] = groupdf.apply(apply2,axis=1,result_type='expand')


    def tests(row):
            res = []
            
            if 0 < row["btlrateCur"] < 0.031 and row["Division"] != 90:
                res.append("שיעור ביטוח לאומי נמוך ממזערי")
            #
            
            if row["taxrateCur"] > 0.4:
                res.append("שיעור מס גבוה")
            #
            
            if row["taxCur"] < -100 and row["taxPrev"] < -100:
                res.append("מס שלילי חודשיים ברציפות")
            #
            
            return ". ".join(res)
        #
        
    groupdf["ErrorDescr"] = groupdf.apply(tests,axis=1)
    
    cols = ["Empid","Empname","CurGross","PrevGross","btlCur","btlPrev","taxCur","taxPrev","btlrateCur","taxrateCur","ErrorDescr"]
    
    wb = load_workbook(filename = custom.xlresfile)
    wb.create_sheet(title="מס חריג")
    ws = wb["מס חריג"]

    ws['A1'] = "מספר עובד"
    ws['B1'] = "שם"
    ws['C1'] = "ברוטו שוטף"
    ws['D1'] = "ברוטו קודם"
    ws['E1'] = "בל שוטף"
    ws['F1'] = "בל קודם"
    ws['G1'] = "מס שוטף"
    ws['H1'] = "מס קודם"
    ws['I1'] = "% בל שוטף"
    ws['J1'] = "% מס שוטף"
    ws['K1'] = "תאור שגיאה"

    wb.save(custom.xlresfile)

    with pd.ExcelWriter(custom.xlresfile, mode="a",if_sheet_exists='overlay') as writer:
        groupdf.loc[groupdf["ErrorDescr"] != "",cols].to_excel(writer,sheet_name="מס חריג",index=False,startrow=1,header=False)
    #  

    conn.close()

    return [inspect.stack()[0][3],len(groupdf.loc[groupdf["ErrorDescr"] != "","Empid"].unique()),"מספר עובדים עם שיעור מס וביטוח לאומי חריגים"]
#