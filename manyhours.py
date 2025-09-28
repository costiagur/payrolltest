#עובדים עם מספר שעות רב מדי
import custom
import pandas as pd
import inspect
from openpyxl import Workbook, load_workbook
import sqlite3


def manyhours(level="264"):
    
    level= float(level)

    conn = sqlite3.connect("dbsave.db")
    cur = conn.cursor()

    REFMONTH = cur.execute("SELECT MAX(Refdate) FROM dfcurr").fetchone()[0]

    query = f"SELECT Empid,mn,Empname,Elem,Elem_heb,Quantity,WorkHours FROM dfcurr WHERE Division <> 90 AND Refdate = '{REFMONTH}' AND Elem IN ({custom.yesod},{custom.konenut})"

    middf = pd.read_sql_query(query, conn)

    conn.close()

    #middf = custom.DFCURR.loc[(custom.DFCURR['Elem'].isin((custom.yesod,custom.konenut)))&(custom.DFCURR["Division"]!=custom.pensiondepartment)&(custom.DFCURR["Refdate"]==custom.REFMONTH),["Empid","Empname","mn","Elem_heb","Quantity","Elem","WorkHours"]]

    middf["NetWorkQ"] = middf.apply(lambda row: -1 * row["Quantity"] if row["Elem"] == custom.konenut else row["WorkHours"],axis=1)

    groupdf = middf.groupby(by = ["Empid","Empname","mn"],as_index=False,group_keys=True).sum("NetWorkQ")

    with pd.ExcelWriter(custom.xlresfile, mode="a") as writer:
        groupdf.loc[groupdf["NetWorkQ"] >=level].to_excel(writer,sheet_name="מספר שעות רב",index=False)
    #
 
    wb = load_workbook(filename = custom.xlresfile)
    ws = wb["מספר שעות רב"]

    ws['A1'] = "מספר_עובד"
    ws['B1'] = "שם"
    ws['C1'] = "מנ"
    ws['D1'] = "כמות"
    ws['E1'] = "שעות_עבודה"
    ws['F1'] = "שעות_נטו"

    wb.save(custom.xlresfile) 
    
    return [inspect.stack()[0][3],groupdf.loc[groupdf["NetWorkQ"] >=level].shape[0],"מספר עובדים עם כמות שעות גבוהה מדי"]
#